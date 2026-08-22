"""Unit tests for Architecture Upgrade Proposal components (FactStore, CrossModalValidator, AbstractBudgetEngine, and Prompts)."""

import json
import os
import tempfile
import unittest

from app.core.prompts.coder import CODER_PROMPT
from app.core.prompts.modeler import MODELER_PROMPT
from app.core.prompts.writer import get_writer_prompt
from app.tools.abstract_budget_engine import AbstractBudgetEngine
from app.tools.cross_modal_validator import (
    audit_cross_modal,
    extract_code_generated_files,
    validate_code_text_parity,
)
from app.tools.fact_store import FactStore


class TestArchitectureUpgradeComponents(unittest.TestCase):
    """验证架构升级方案落地模块的功能。"""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.work_dir = self.temp_dir.name

    def tearDown(self):
        self.temp_dir.cleanup()

    # ------------------ 1. FactStore 单元测试 ------------------
    def test_fact_store_register_and_lookup(self):
        store = FactStore(work_dir=self.work_dir)
        store.register_metric(
            name="phi_star",
            value=0.013487,
            subtask_id="ques3",
            unit="%",
            label="最优体积分数",
            source="ques3_results.csv",
            wilson_low=0.8564,
            wilson_high=0.9383,
        )

        fact = store.get("phi_star", subtask_id="ques3")
        self.assertIsNotNone(fact)
        self.assertEqual(fact.value, 0.013487)
        self.assertEqual(fact.format_value(precision=6), "0.013487")
        self.assertEqual(store.get_value("phi_star", subtask_id="ques3"), 0.013487)

    def test_fact_store_placeholder_rendering(self):
        store = FactStore(work_dir=self.work_dir)
        store.register_metric(
            name="cost_star",
            value=7.88,
            subtask_id="ques4",
            unit="元",
            label="最低成本",
            source="ques4_optimal_solution.csv",
        )
        store.register_metric(
            name="na_star",
            value=531,
            subtask_id="ques4",
            unit="根",
            label="最优圆柱数量",
            source="ques4_optimal_solution.csv",
        )

        template = (
            "求得最优决策为 N_A = {{ facts.ques4.na_star }} 根，"
            "最低总成本为 {{ facts.ques4.cost_star }} 元。"
        )
        rendered, replacements = store.render_template(template)
        self.assertIn("N_A = 531 根", rendered)
        self.assertIn("最低总成本为 7.88 元", rendered)
        self.assertEqual(len(replacements), 2)

    def test_fact_store_disk_persistence_and_reloading(self):
        store = FactStore(work_dir=self.work_dir)
        store.register_metric(name="p_target", value=0.90, subtask_id="ques3")
        saved_path = store.save_to_disk(self.work_dir)
        self.assertTrue(os.path.isfile(saved_path))

        reloaded = FactStore.load_from_disk(self.work_dir)
        self.assertEqual(reloaded.get_value("p_target", subtask_id="ques3"), 0.90)

    # ------------------ 2. CrossModalValidator 单元测试 ------------------
    def test_ast_file_generation_extraction(self):
        sample_code = """
import pandas as pd
import numpy as np

df = pd.DataFrame({"col": [1, 2, 3]})
df.to_csv("ques4_global_frontier_certificate.csv", index=False)
df.to_excel("ques4_results.xlsx")

with open("ques3_acceptance_metrics.csv", "w") as f:
    f.write("a,b\\n1,2\\n")
"""
        generated_files, output_calls = extract_code_generated_files(sample_code)
        self.assertIn("ques4_global_frontier_certificate.csv", generated_files)
        self.assertIn("ques4_results.xlsx", generated_files)
        self.assertIn("ques3_acceptance_metrics.csv", generated_files)
        self.assertGreaterEqual(len(output_calls), 3)

    def test_code_text_parity_detection(self):
        markdown_text = """
# 五、模型求解与分析
我们在研究中通过计算生成了全域排除证书 `ques4_global_frontier_certificate.csv`，
并且所有数据已记录在 `ques3_results.csv` 中。
"""
        # 场景 A: 代码中完整包含了两个文件的生成
        code_complete = """
df1.to_csv("ques4_global_frontier_certificate.csv")
df2.to_csv("ques3_results.csv")
"""
        res_pass = validate_code_text_parity(markdown_text, code_complete, self.work_dir)
        self.assertTrue(res_pass["passed"])
        self.assertEqual(res_pass["status"], "PASS")
        self.assertEqual(len(res_pass["missing_critical_generators"]), 0)

        # 场景 B: 正文声称有证书，但代码中缺失
        code_incomplete = """
df2.to_csv("ques3_results.csv")
"""
        res_fail = validate_code_text_parity(markdown_text, code_incomplete, self.work_dir)
        self.assertFalse(res_fail["passed"])
        self.assertIn("ques4_global_frontier_certificate.csv", res_fail["missing_critical_generators"])

    def test_audit_cross_modal_writes_report(self):
        report = audit_cross_modal(self.work_dir, markdown_text="正文测试", code_sources=[])
        report_path = os.path.join(self.work_dir, "cross_modal_audit.json")
        self.assertTrue(os.path.isfile(report_path))
        self.assertIn("status", report)

    def test_audit_optimality_certificates_catches_contradictions(self):
        from app.tools.cross_modal_validator import audit_optimality_certificates

        # 创建一个宣称最优成本为 7.88 的最优解文件
        opt_path = os.path.join(self.work_dir, "ques4_optimal_solution.csv")
        with open(opt_path, "w", encoding="utf-8") as f:
            f.write("total_cost_yuan,N_A,N_B\n7.882177,531,0\n")

        # 场景 A: 证书中的点全部被成功排除 (Wilson_low < 0.90)
        cert_pass_path = os.path.join(self.work_dir, "ques4_global_frontier_certificate.csv")
        with open(cert_pass_path, "w", encoding="utf-8") as f:
            f.write("N_A,N_B_max,total_cost_yuan,wilson_low,status\n")
            f.write("500,274,7.881104,0.855150,EXCLUDED\n")
            f.write("520,97,7.881418,0.888400,EXCLUDED\n")

        res_pass = audit_optimality_certificates(self.work_dir)
        self.assertTrue(res_pass["passed"])
        self.assertEqual(res_pass["status"], "PASS")
        self.assertEqual(res_pass["contradiction_count"], 0)

        # 场景 B: 证书中存在一个成本更低但被标记为 FEASIBLE 的矛盾点
        with open(cert_pass_path, "w", encoding="utf-8") as f:
            f.write("N_A,N_B_max,total_cost_yuan,wilson_low,status\n")
            f.write("500,274,7.881104,0.855150,EXCLUDED\n")
            f.write("530,8,7.880737,0.901931,FEASIBLE\n")

        res_fail = audit_optimality_certificates(self.work_dir)
        self.assertFalse(res_fail["passed"])
        self.assertEqual(res_fail["status"], "FAIL")
        self.assertEqual(res_fail["contradiction_count"], 1)
        self.assertIn("530", str(res_fail["contradictions"]))

    # ------------------ 3. AbstractBudgetEngine 单元测试 ------------------
    def test_abstract_budget_engine_micro_adjustments(self):
        adjustments_mod = AbstractBudgetEngine.get_adaptive_micro_adjustments("moderate")
        self.assertIn("linestretch", adjustments_mod)
        self.assertIn("fontsize", adjustments_mod)

        adjustments_heavy = AbstractBudgetEngine.get_adaptive_micro_adjustments("heavy")
        self.assertEqual(adjustments_heavy["linestretch"], "1.15")

    # ------------------ 4. 提示词升级契约测试 ------------------
    def test_modeler_prompt_contains_rigor_and_literature(self):
        self.assertIn("Wilson 95% 置信区间", MODELER_PROMPT)
        self.assertIn("一维上确界前沿排除证明算法", MODELER_PROMPT)
        self.assertIn("模型选型锦标赛与备选方案证伪", MODELER_PROMPT)
        self.assertIn("假设与解释标准化规范", MODELER_PROMPT)
        self.assertIn("可靠性约束优化模型", MODELER_PROMPT)

    def test_coder_prompt_contains_advanced_templates(self):
        self.assertIn("Composite Figures", CODER_PROMPT)
        self.assertIn("Mechanism Diagrams", CODER_PROMPT)
        self.assertIn("wilson_score_interval", CODER_PROMPT)
        self.assertIn("SELF-CONTAINED SOLVER PROTOCOL", CODER_PROMPT)
        self.assertIn("GENTLE CONCURRENCY PROTOCOL", CODER_PROMPT)

    def test_writer_prompt_contains_consistency_and_literature(self):
        writer_prompt = get_writer_prompt()
        self.assertIn("复合多子图解析", writer_prompt)
        self.assertIn("机理图/算法流程图解读", writer_prompt)
        self.assertIn("统计标准统一性", writer_prompt)
        self.assertIn("关键词数模术语硬性规范", writer_prompt)
        self.assertIn("可靠性约束优化模型", writer_prompt)

    def test_audit_latex_formatting_integrity(self):
        from app.tools.cross_modal_validator import audit_latex_formatting_integrity

        # 正确格式
        clean_text = "在本文最优解 $M=5000$ 密集采样下，求得最优解为 $N_A^*=531, N_B^*=0, C^*=7.88\\text{ 元}$。"
        res_clean = audit_latex_formatting_integrity(clean_text)
        self.assertTrue(res_clean["passed"])
        self.assertEqual(res_clean["issue_count"], 0)

        # 损坏格式（正文变量展开损坏，如 =500$, .90$, .88\text）
        corrupt_text = "基准样本量 =500$ 且概率低于 .90$，总成本低于 .88\\text{ 元}$。"
        res_corrupt = audit_latex_formatting_integrity(corrupt_text)
        self.assertFalse(res_corrupt["passed"])
        self.assertGreater(res_corrupt["issue_count"], 0)

        # 附录代码块内注释损坏检测
        code_fence_corrupt = (
            "### 附录B 源码\n"
            "```python\n"
            "# 经扫描（基准样本量 =500$ 且概率严格低于 .90$）\n"
            "def solver():\n"
            "    pass\n"
            "```\n"
        )
        res_code_corrupt = audit_latex_formatting_integrity(code_fence_corrupt)
        self.assertFalse(res_code_corrupt["passed"])
        self.assertGreater(res_code_corrupt["issue_count"], 0)
        self.assertEqual(res_code_corrupt["issues"][0]["type"], "corrupted_latex_in_code")

    def test_audit_on_fixture_artifacts(self):
        """默认单元测试：始终使用临时目录与固定 fixture，确保 CI 与本地确定性一致。"""
        from app.tools.cross_modal_validator import (
            audit_cross_modal,
            audit_latex_formatting_integrity,
        )

        res_md_fixture = (
            "# 题目摘要\n\n## 摘要\n本文基于微构体模型建立可靠性约束优化模型。\n\n"
            "关键词：混合整数规划；自适应抽样；灵敏度分析\n\n"
            "# 五、模型的建立与求解\n"
            "通过计算生成了全域排除证书 `ques4_global_frontier_certificate.csv`。\n\n"
            "# 附录\n## 附录B 源程序代码\n"
            "```python\n# 独立求解器\nimport numpy as np\nimport pandas as pd\n"
            "df = pd.DataFrame({'N_A': [531]})\ndf.to_csv('ques4_global_frontier_certificate.csv')\n```\n"
        )
        with open(os.path.join(self.work_dir, "res.md"), "w", encoding="utf-8") as f:
            f.write(res_md_fixture)
        with open(os.path.join(self.work_dir, "master_solver.py"), "w", encoding="utf-8") as f:
            f.write("import numpy as np\nimport pandas as pd\n")
        with open(os.path.join(self.work_dir, "ques4_global_frontier_certificate.csv"), "w", encoding="utf-8") as f:
            f.write("N_A,status,total_cost_yuan,wilson_low\n500,EXCLUDED,7.5,0.85\n")

        res_md_path = os.path.join(self.work_dir, "res.md")
        with open(res_md_path, encoding="utf-8", errors="replace") as f:
            md_content = f.read()

        # 1. LaTeX 与代码块格式门禁
        latex_res = audit_latex_formatting_integrity(md_content)
        self.assertIn("passed", latex_res)
        self.assertTrue(
            latex_res["passed"],
            f"res.md 未通过 LaTeX 完整性审计: {latex_res.get('issues')}",
        )

        # 2. 跨模态完整审计
        cross_res = audit_cross_modal(self.work_dir)
        self.assertIn("status", cross_res)
        self.assertTrue(
            cross_res["passed"],
            f"跨模态审计未通过: {cross_res}",
        )
        self.assertIn(
            cross_res["status"],
            ("PASS", "WARN"),
            f"跨模态审计状态异常: {cross_res}",
        )

    def test_audit_on_explicit_integration_task_if_configured(self):
        """显式集成测试：仅当环境变量提供 MMA_INTEGRATION_TASK_DIR 且路径存在时才执行真实任务目录核验。"""
        from app.tools.cross_modal_validator import (
            audit_cross_modal,
            audit_latex_formatting_integrity,
        )

        integration_task_dir = os.environ.get("MMA_INTEGRATION_TASK_DIR")
        if not integration_task_dir or not os.path.isdir(integration_task_dir):
            self.skipTest("未配置 MMA_INTEGRATION_TASK_DIR，跳过真实任务集成测试")

        res_md_path = os.path.join(integration_task_dir, "res.md")
        if not os.path.isfile(res_md_path):
            self.skipTest("指定任务目录下不存在 res.md，跳过真实任务集成测试")

        with open(res_md_path, encoding="utf-8", errors="replace") as f:
            md_content = f.read()

        latex_res = audit_latex_formatting_integrity(md_content)
        self.assertTrue(latex_res["passed"])

        cross_res = audit_cross_modal(integration_task_dir)
        self.assertTrue(cross_res["passed"])

    def test_audit_code_self_containment(self):
        from app.tools.cross_modal_validator import audit_code_self_containment

        # 1. 干净自包含代码（PASS）
        clean_code = """
import numpy as np
import pandas as pd

class UniformGridBroadphase3DPBC:
    def __init__(self, box_size, cell_size):
        self.box_size = box_size
"""
        res_clean = audit_code_self_containment(code_sources=[{"name": "master_solver.py", "code": clean_code}])
        self.assertTrue(res_clean["passed"])
        self.assertEqual(res_clean["status"], "PASS")
        self.assertEqual(res_clean["issue_count"], 0)

        # 2. 包含仓库私有导入（FAIL）
        corrupt_code_import = """
import numpy as np
from app.tools.geometric_lib import UniformGridBroadphase3DPBC
"""
        res_import = audit_code_self_containment(code_sources=[{"name": "solver.py", "code": corrupt_code_import}])
        self.assertFalse(res_import["passed"])
        self.assertEqual(res_import["status"], "FAIL")
        self.assertGreaterEqual(res_import["issue_count"], 1)
        self.assertIn("app.tools.geometric_lib", str(res_import["issues"]))

        # 3. 包含 sys.path 追加（FAIL）
        corrupt_code_sys = """
import sys
sys.path.append("D:/workspace/MathModelAgent/backend")
"""
        res_sys = audit_code_self_containment(code_sources=[{"name": "solver.py", "code": corrupt_code_sys}])
        self.assertFalse(res_sys["passed"])
        self.assertEqual(res_sys["status"], "FAIL")
        self.assertGreaterEqual(res_sys["issue_count"], 1)

        # 4. 工作目录多求解脚本扫描（覆盖 ques1_solver.py 等）
        with open(os.path.join(self.work_dir, "ques1_solver.py"), "w", encoding="utf-8") as f:
            f.write("from app.utils.log_util import logger\n")
        res_workdir = audit_code_self_containment(work_dir=self.work_dir)
        self.assertFalse(res_workdir["passed"])
        self.assertIn("ques1_solver.py", str(res_workdir["issues"]))

    def test_refresh_frozen_results_hashes_with_equivalence_and_conflict(self):
        import hashlib
        import json
        from app.tools.result_integrity import refresh_frozen_results_hashes, validate_result_freeze

        csv_path = os.path.join(self.work_dir, "ques1_results.csv")
        with open(csv_path, "w", encoding="utf-8", newline="\n") as f:
            f.write("metric,value\nx_val,10\n")

        with open(csv_path, "rb") as f:
            initial_hash = hashlib.sha256(f.read()).hexdigest()

        frozen_doc = {
            "schema": "mathmodel.result-freeze",
            "version": 1,
            "metrics": [
                {
                    "id": "x_val",
                    "label": "X指标",
                    "value": 10,
                    "unit": "个",
                    "explanation": "测试指标",
                    "source_path": "ques1_results.csv",
                }
            ],
            "sources": [
                {
                    "relative_path": "ques1_results.csv",
                    "sha256": initial_hash,
                }
            ],
        }
        frozen_path = os.path.join(self.work_dir, "frozen_results.json")
        with open(frozen_path, "w", encoding="utf-8") as f:
            json.dump(frozen_doc, f, indent=2)

        # 验证初始状态通过
        val_init = validate_result_freeze(self.work_dir)
        self.assertTrue(val_init["passed"])

        # 场景 A: 格式/换行调整但数值等价 (x_val 仍为 10)
        with open(csv_path, "w", encoding="utf-8", newline="\r\n") as f:
            f.write("metric,value\r\nx_val, 10.000\r\n")

        # 此时因字节改变 validate 失败
        self.assertFalse(validate_result_freeze(self.work_dir)["passed"])

        # 等价性刷新应当成功放行
        refresh_equiv = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_equiv["active"])
        self.assertTrue(refresh_equiv["updated"])
        self.assertFalse(refresh_equiv["has_conflicts"])
        self.assertTrue(validate_result_freeze(self.work_dir)["passed"])

        # 场景 B: 数值发生实质性突变 (x_val 变成 20) -> 必须拦截！
        with open(csv_path, "w", encoding="utf-8", newline="\n") as f:
            f.write("metric,value\nx_val,20\n")

        refresh_conflict = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_conflict["has_conflicts"])
        self.assertEqual(len(refresh_conflict["conflicts"]), 1)
        self.assertIn("已变更为 20.0", refresh_conflict["conflicts"][0]["reason"])

        # 哈希未被偷换，validate_result_freeze 保持报错拦截
        self.assertFalse(validate_result_freeze(self.work_dir)["passed"])

    def test_refresh_frozen_results_hashes_xlsx_and_json_equivalence_and_conflict(self):
        import hashlib
        import json
        import openpyxl
        from app.tools.result_integrity import refresh_frozen_results_hashes, validate_result_freeze

        # 1. 测试 XLSX 数据源
        xlsx_path = os.path.join(self.work_dir, "ques2_results.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["id", "label", "value"])
        ws.append(["optimal_cost", "最优成本", 7.882])
        wb.save(xlsx_path)
        wb.close()

        with open(xlsx_path, "rb") as f:
            initial_xlsx_hash = hashlib.sha256(f.read()).hexdigest()

        frozen_doc = {
            "schema": "mathmodel.result-freeze",
            "version": 1,
            "metrics": [
                {
                    "id": "optimal_cost",
                    "label": "最优成本",
                    "value": 7.882,
                    "unit": "元",
                    "explanation": "最小化总成本",
                    "source_path": "ques2_results.xlsx",
                }
            ],
            "sources": [
                {
                    "relative_path": "ques2_results.xlsx",
                    "sha256": initial_xlsx_hash,
                }
            ],
        }
        frozen_path = os.path.join(self.work_dir, "frozen_results.json")
        with open(frozen_path, "w", encoding="utf-8") as f:
            json.dump(frozen_doc, f, indent=2)

        self.assertTrue(validate_result_freeze(self.work_dir)["passed"])

        # 重新保存 XLSX（格式调整但数值等价: 7.8820 -> sha256 变动但数值等价）
        wb2 = openpyxl.load_workbook(xlsx_path)
        ws2 = wb2.active
        ws2["C2"] = "7.8820"
        wb2.save(xlsx_path)
        wb2.close()

        # 此时因 zip 字节变动导致校验失败
        self.assertFalse(validate_result_freeze(self.work_dir)["passed"])

        # 等价性刷新应当成功放行并更新哈希
        refresh_equiv = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_equiv["active"])
        self.assertTrue(refresh_equiv["updated"])
        self.assertFalse(refresh_equiv["has_conflicts"])
        self.assertTrue(validate_result_freeze(self.work_dir)["passed"])

        # 修改 XLSX 数值为 9.999 -> 必须拦截！
        wb3 = openpyxl.Workbook()
        ws3 = wb3.active
        ws3.append(["id", "label", "value"])
        ws3.append(["optimal_cost", "最优成本", 9.999])
        wb3.save(xlsx_path)
        wb3.close()

        refresh_conflict = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_conflict["has_conflicts"])
        self.assertEqual(len(refresh_conflict["conflicts"]), 1)
        self.assertIn("已变更为 9.999", refresh_conflict["conflicts"][0]["reason"])

        # 2. 测试 JSON 数据源
        json_src_path = os.path.join(self.work_dir, "ques3_results.json")
        with open(json_src_path, "w", encoding="utf-8") as f:
            json.dump([{"id": "prob", "value": 0.9122}], f)

        with open(json_src_path, "rb") as f:
            initial_json_hash = hashlib.sha256(f.read()).hexdigest()

        frozen_doc["metrics"] = [
            {
                "id": "prob",
                "label": "导通概率",
                "value": 0.9122,
                "unit": "%",
                "explanation": "蒙特卡洛导通概率",
                "source_path": "ques3_results.json",
            }
        ]
        frozen_doc["sources"] = [
            {
                "relative_path": "ques3_results.json",
                "sha256": initial_json_hash,
            }
        ]
        with open(frozen_path, "w", encoding="utf-8") as f:
            json.dump(frozen_doc, f, indent=2)

        # 格式化 JSON（字节改变但数值等价）
        with open(json_src_path, "w", encoding="utf-8") as f:
            json.dump([{"id": "prob", "value": 0.9122}], f, indent=4)

        # 字节改变导致校验失败
        self.assertFalse(validate_result_freeze(self.work_dir)["passed"])

        refresh_json_equiv = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_json_equiv["updated"])
        self.assertFalse(refresh_json_equiv["has_conflicts"])
        self.assertTrue(validate_result_freeze(self.work_dir)["passed"])

        # 修改 JSON 数值为 0.5000 -> 必须拦截！
        with open(json_src_path, "w", encoding="utf-8") as f:
            json.dump([{"id": "prob", "value": 0.5000}], f)

        refresh_json_conflict = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_json_conflict["has_conflicts"])
        self.assertIn("已变更为 0.5", refresh_json_conflict["conflicts"][0]["reason"])

        # 3. 测试 XLSX 多 Sheet 场景（指标分布在非首个 Sheet）
        multi_sheet_path = os.path.join(self.work_dir, "ques4_multi_sheet.xlsx")
        wb_m = openpyxl.Workbook()
        ws_m1 = wb_m.active
        ws_m1.title = "Cover"
        ws_m1.append(["Title", "Description"])
        ws_m1.append(["Experiment 4", "Summary sheet"])
        ws_m2 = wb_m.create_sheet(title="MetricsData")
        ws_m2.append(["id", "label", "value"])
        ws_m2.append(["percolation_threshold", "渗流阈值", 0.4215])
        wb_m.save(multi_sheet_path)
        wb_m.close()

        with open(multi_sheet_path, "rb") as f:
            multi_sheet_hash = hashlib.sha256(f.read()).hexdigest()

        frozen_doc["metrics"] = [
            {
                "id": "percolation_threshold",
                "label": "渗流阈值",
                "value": 0.4215,
                "unit": "无量纲",
                "explanation": "渗流体积分数阈值",
                "source_path": "ques4_multi_sheet.xlsx",
            }
        ]
        frozen_doc["sources"] = [
            {
                "relative_path": "ques4_multi_sheet.xlsx",
                "sha256": multi_sheet_hash,
            }
        ]
        with open(frozen_path, "w", encoding="utf-8") as f:
            json.dump(frozen_doc, f, indent=2)

        self.assertTrue(validate_result_freeze(self.work_dir)["passed"])

        # 修改 Sheet2 中的数值为 0.8888 -> 即使在非首个 Sheet 也必须被精准拦截！
        wb_m3 = openpyxl.load_workbook(multi_sheet_path)
        ws_m2_mod = wb_m3["MetricsData"]
        ws_m2_mod["C2"] = 0.8888
        wb_m3.save(multi_sheet_path)
        wb_m3.close()

        refresh_multi_conflict = refresh_frozen_results_hashes(self.work_dir, verify_equivalence=True)
        self.assertTrue(refresh_multi_conflict["has_conflicts"])
        self.assertIn("已变更为 0.8888", refresh_multi_conflict["conflicts"][0]["reason"])

    def test_prompts_precision_and_concurrency_policy(self):
        # 1. Writer Prompt
        writer_prompt = get_writer_prompt()
        self.assertIn("可靠性约束优化模型", writer_prompt)
        self.assertIn("Wilson 95% 置信下限准入规则", writer_prompt)
        self.assertIn("基于临界加密的分层自适应 Monte Carlo 检验", writer_prompt)
        self.assertIn("核心单文件自包含求解器", writer_prompt)
        self.assertIn("关键词数模术语硬性规范", writer_prompt)
        self.assertIn("假设 1：[假设内容简练陈述]。解释：", writer_prompt)

        # 2. Modeler Prompt
        self.assertIn("可靠性约束优化模型", MODELER_PROMPT)
        self.assertIn("分层自适应两阶段 Monte Carlo 检验", MODELER_PROMPT)
        self.assertIn("准入决策金句标准范式", MODELER_PROMPT)
        self.assertIn("假设与解释标准化规范", MODELER_PROMPT)
        self.assertIn("模型选型锦标赛与备选方案证伪", MODELER_PROMPT)

        # 3. Coder Prompt
        self.assertIn("SELF-CONTAINED SOLVER PROTOCOL", CODER_PROMPT)
        self.assertIn("GENTLE CONCURRENCY PROTOCOL", CODER_PROMPT)
        self.assertIn("ProcessPoolExecutor", CODER_PROMPT)
        self.assertIn("分层自适应两阶段抽样", CODER_PROMPT)

    def test_audit_keywords_modeling_compliance(self):
        from app.tools.cross_modal_validator import audit_keywords_modeling_compliance

        # 1. 合规数模关键词
        valid_md = """
# 论文标题
## 摘要
本文建立了混合整数非线性规划模型。
关键词：混合整数规划；蒙特卡洛模拟；自适应抽样；灵敏度分析
# 一、问题重述
"""
        res_valid = audit_keywords_modeling_compliance(valid_md)
        self.assertTrue(res_valid["passed"])
        self.assertEqual(res_valid["status"], "PASS")
        self.assertEqual(len(res_valid["keywords"]), 4)

        # 2. 纯背景词（违规预警提示：status=WARN, passed=True）
        invalid_md = """
# 论文标题
## 摘要
本文研究了乡村农业与蔬菜销售。
关键词：乡村；农业；蔬菜；种植策略
# 一、问题重述
"""
        res_invalid = audit_keywords_modeling_compliance(invalid_md)
        self.assertTrue(res_invalid["passed"])
        self.assertEqual(res_invalid["status"], "WARN")
        self.assertIn("domain_only_keywords", [i["type"] for i in res_invalid["issues"]])

        # 3. 领域词仅包含“优化”子串不应直接放行
        domain_opt_md = """
# 论文标题
## 摘要
关键词：种植优化；农作物；销售策略
# 一、问题重述
"""
        res_domain_opt = audit_keywords_modeling_compliance(domain_opt_md)
        self.assertEqual(res_domain_opt["status"], "WARN")
        self.assertTrue(res_domain_opt["passed"])
        self.assertGreater(len(res_domain_opt["issues"]), 0)

    def test_cross_modal_status_invariants(self):
        """测试跨模态审计的状态不变量：阻断项必为 FAIL/False，仅预警项必为 WARN/True。"""
        # 场景 A: 仅有关键词预警 -> status="WARN", passed=True
        warn_md = """
# 论文标题
## 摘要
关键词：乡村；农业；蔬菜
# 一、问题重述
正文。
"""
        res_warn = audit_cross_modal(self.work_dir, markdown_text=warn_md, code_sources=[])
        self.assertEqual(res_warn["status"], "WARN")
        self.assertTrue(res_warn["passed"])

        # 场景 B: 代码包含私有依赖阻断项 -> status="FAIL", passed=False
        bad_code = "from app.utils.log_util import logger\n"
        res_fail = audit_cross_modal(
            self.work_dir,
            markdown_text=warn_md,
            code_sources=[{"name": "solver.py", "code": bad_code}],
        )
        self.assertEqual(res_fail["status"], "FAIL")
        self.assertFalse(res_fail["passed"])

    def test_formal_solver_syntax_error_and_out_of_bounds(self):
        """测试正式求解器语法错误与越界路径检测。"""
        from app.tools.cross_modal_validator import audit_code_self_containment

        # 1. 求解器语法错误
        syntax_err_path = os.path.join(self.work_dir, "master_solver.py")
        with open(syntax_err_path, "w", encoding="utf-8") as f:
            f.write("def broken_syntax(:\n    pass\n")

        res_syntax = audit_code_self_containment(work_dir=self.work_dir)
        self.assertFalse(res_syntax["passed"])
        self.assertEqual(res_syntax["status"], "FAIL")
        self.assertTrue(any(i["type"] == "formal_solver_syntax_error" for i in res_syntax["issues"]))

        # 2. 越界代码源路径
        os.remove(syntax_err_path)
        frozen_doc = {
            "schema": "mathmodel.result-freeze",
            "version": 1,
            "metrics": [],
            "sources": [],
            "executed_code_sources": ["../outside_solver.py"],
        }
        with open(os.path.join(self.work_dir, "frozen_results.json"), "w", encoding="utf-8") as f:
            json.dump(frozen_doc, f)

        res_oob = audit_code_self_containment(work_dir=self.work_dir)
        self.assertFalse(res_oob["passed"])
        self.assertEqual(res_oob["status"], "FAIL")
        self.assertTrue(any(i["type"] == "out_of_bounds_source_path" for i in res_oob["issues"]))


if __name__ == "__main__":
    unittest.main()

"""Structured result-freeze helpers used by the writer and paper preflight.

The module deliberately does not infer mathematical facts from prose.  A
``frozen_results.json`` file is an explicit hand-off from a successful code
execution/validation step to the writer.  Its numbers are therefore the only
computed numbers that a writer may use once the file is present.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
from typing import Any


FREEZE_FILENAMES = (
    "frozen_results.json",
    "result_freeze.json",
    "reports/frozen_numbers.json",  # compatible with skill 3a-result-freeze
)
FREEZE_SCHEMAS = {"mathmodel.result-freeze", "mathmodel.writer-result-freeze"}
FREEZE_VERSION = 1
_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_path(work_dir: str, relative_path: str) -> str | None:
    if not isinstance(relative_path, str) or not relative_path.strip():
        return None
    root = os.path.realpath(work_dir)
    candidate = os.path.realpath(os.path.join(root, relative_path))
    try:
        if os.path.commonpath((root, candidate)) != root:
            return None
    except ValueError:
        return None
    return candidate


def load_result_freeze(work_dir: str) -> dict[str, Any] | None:
    """Return the first valid-looking freeze document, or ``None`` if absent.

    The returned value includes only a relative ``_path`` helper.  Detailed
    schema errors are reported by :func:`validate_result_freeze`, so callers
    can distinguish a missing optional artifact from an invalid active one.
    """
    for relative_path in FREEZE_FILENAMES:
        path = os.path.join(work_dir, relative_path)
        if not os.path.isfile(path):
            continue
        try:
            with open(path, encoding="utf-8") as handle:
                document = json.load(handle)
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            return {"_path": relative_path, "_invalid_json": True}
        if not isinstance(document, dict):
            return {"_path": relative_path, "_invalid_json": True}
        document = dict(document)
        document["_path"] = relative_path.replace("\\", "/")
        return document
    return None


def _metric_label(metric: dict[str, Any]) -> str:
    return str(metric.get("label") or metric.get("id") or "").strip()


def _metric_aliases(metric: dict[str, Any]) -> list[str]:
    aliases = [_metric_label(metric)]
    raw_aliases = metric.get("aliases", [])
    if isinstance(raw_aliases, list):
        aliases.extend(str(item).strip() for item in raw_aliases)
    return [item for item in dict.fromkeys(aliases) if item]


def validate_result_freeze(work_dir: str, document: dict[str, Any] | None = None) -> dict[str, Any]:
    """Validate schema, metric semantics and evidence hashes for a freeze."""
    document = document if document is not None else load_result_freeze(work_dir)
    if document is None:
        return {"active": False, "passed": True, "errors": [], "metrics": []}

    errors: list[dict[str, str]] = []
    if document.get("_invalid_json"):
        return {
            "active": True,
            "passed": False,
            "path": document.get("_path", "frozen_results.json"),
            "errors": [{"code": "freeze_unreadable", "detail": "冻结结果文件不是有效 JSON。"}],
            "metrics": [],
        }
    if document.get("schema") not in FREEZE_SCHEMAS or document.get("version") != FREEZE_VERSION:
        errors.append({"code": "freeze_schema_invalid", "detail": "冻结结果 schema 或版本无效。"})

    raw_metrics = document.get("metrics")
    metrics: list[dict[str, Any]] = []
    if not isinstance(raw_metrics, list) or not raw_metrics:
        errors.append({"code": "metrics_missing", "detail": "冻结结果必须含有非空 metrics。"})
    else:
        metric_ids: set[str] = set()
        for index, raw_metric in enumerate(raw_metrics):
            if not isinstance(raw_metric, dict):
                errors.append({"code": "metric_invalid", "detail": f"metrics[{index}] 不是对象。"})
                continue
            metric = dict(raw_metric)
            metric_id = str(metric.get("id", "")).strip()
            label = _metric_label(metric)
            if not metric_id or not label:
                errors.append({"code": "metric_identity_missing", "detail": f"metrics[{index}] 缺少 id 或 label。"})
            elif metric_id in metric_ids:
                errors.append({"code": "metric_duplicate", "detail": f"重复指标 id：{metric_id}。"})
            metric_ids.add(metric_id)
            if not isinstance(metric.get("value"), (int, float)) or isinstance(metric.get("value"), bool):
                errors.append({"code": "metric_value_invalid", "detail": f"指标 {metric_id or index} 的 value 必须是数值。"})
            if not str(metric.get("unit", "")).strip() or not str(metric.get("explanation", "")).strip():
                errors.append({"code": "metric_semantics_missing", "detail": f"指标 {metric_id or index} 缺少 unit 或 explanation。"})
            metrics.append(metric)

    sources = document.get("sources", [])
    if not isinstance(sources, list) or not sources:
        errors.append({"code": "sources_missing", "detail": "冻结结果必须包含可核验的来源文件。"})
    else:
        for index, source in enumerate(sources):
            if not isinstance(source, dict):
                errors.append({"code": "source_invalid", "detail": f"sources[{index}] 不是对象。"})
                continue
            relative_path = source.get("relative_path") or source.get("path")
            expected_hash = source.get("sha256")
            source_path = _safe_path(work_dir, str(relative_path or ""))
            if source_path is None:
                errors.append({"code": "source_path_invalid", "detail": f"sources[{index}] 路径无效。"})
                continue
            if not isinstance(expected_hash, str) or not _SHA256_RE.fullmatch(expected_hash):
                errors.append({"code": "source_hash_invalid", "detail": f"来源 {relative_path} 缺少有效 SHA-256。"})
                continue
            if not os.path.isfile(source_path):
                errors.append({"code": "source_missing", "detail": f"来源文件不存在：{relative_path}。"})
            elif _sha256(source_path) != expected_hash:
                errors.append({"code": "source_hash_changed", "detail": f"来源文件已变化：{relative_path}。"})

    return {
        "active": True,
        "passed": not errors,
        "path": document.get("_path", "frozen_results.json"),
        "errors": errors,
        "metrics": metrics,
        "document": document,
    }


def build_frozen_result_summary(work_dir: str, subtask_id: str | None = None) -> str:
    """Create the writer-only fact block from a validated result freeze.

    当传入 ``subtask_id`` 时，只输出该子任务自己的冻结指标（物理过滤），用于
    Writer 分节写作的子任务隔离：写 quesN 时看不到其它子任务的 frozen 指标。
    未指定 subtask_id（如 eda/敏感性/全局校验）时返回全部指标，保持原行为。
    """
    validation = validate_result_freeze(work_dir)
    if not validation["active"]:
        return ""
    if not validation["passed"]:
        details = "；".join(error["detail"] for error in validation["errors"])
        return (
            "【冻结结果状态：不可用】\n"
            f"{details}\n"
            "不得写入任何新的计算结论、最优参数或算法效果；应如实说明结果尚未通过可追溯性核验。"
        )

    document = validation["document"]
    target = str(subtask_id).lower() if subtask_id else None

    def _in_scope(metric: dict[str, Any]) -> bool:
        # 只排除“明确归属到其它子任务”的指标；无 subtask_id 的指标无法归属，
        # 放行（与非冻结 CSV 路径一致，避免误删无法归属的合法事实）。
        if target is None:
            return True
        own = str(metric.get("subtask_id", "")).lower()
        if not own:
            return True
        return own == target

    scoped_header = (
        f"【冻结结果事实（唯一数值来源，仅限本题 {subtask_id}）】"
        if target is not None
        else "【冻结结果事实（唯一数值来源）】"
    )
    lines = [
        scoped_header,
        f"冻结文件：{validation['path']}。正文、摘要、图题、结论中的计算结果只能使用下列指标；"
        "不得以代码手自然语言总结、图像目测或模型记忆补写其他数值。题面给定常量须明确标为题设，"
        "不能伪装成计算结果。",
    ]
    for metric in validation["metrics"]:
        if not _in_scope(metric):
            continue
        lines.append(
            f"- {metric['id']}：{_metric_label(metric)} = {metric['value']} {metric['unit']}"
            f"（{metric['explanation']}）"
        )
    for subtask in document.get("subtasks", []):
        if not isinstance(subtask, dict):
            continue
        if target is not None and str(subtask.get("id", "")).lower() != target:
            continue
        feasible = subtask.get("feasible")
        if feasible is False:
            identifier = subtask.get("id") or subtask.get("problem") or "当前子问题"
            lines.append(
                f"- {identifier}：当前执行结果不可行；禁止称其为最优方案、最优解或已满足目标。"
            )
    return "\n".join(lines)


def metric_aliases(metric: dict[str, Any]) -> list[str]:
    """Public, deterministic aliases used by the prose and figure checks."""
    return _metric_aliases(metric)


def _extract_rows_from_source(
    source_path: str,
) -> tuple[list[dict[str, Any]] | None, str | None]:
    """从 CSV、Excel (XLSX/XLS) 或 JSON 数据源提取结构化行记录列表（含 Sheet 与嵌套结构支持）。"""
    ext = os.path.splitext(source_path)[1].lower()
    if ext == ".csv":
        try:
            with open(source_path, encoding="utf-8", errors="replace") as f:
                content = f.read()
            if not content.strip():
                return None, "CSV 数据源文件为空"
            reader = csv.DictReader(content.splitlines())
            rows = list(reader)
            if not rows:
                return None, "CSV 数据源无有效数据行"
            return rows, None
        except Exception as exc:
            return None, f"无法解析 CSV 数据源: {exc}"

    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
            rows: list[dict[str, Any]] = []
            for sheet in wb.worksheets:
                sheet_name = str(sheet.title)
                iter_rows = sheet.iter_rows(values_only=True)
                header_row = next(iter_rows, None)
                if not header_row:
                    continue
                headers = [str(h).strip() if h is not None else "" for h in header_row]
                for row_idx, r in enumerate(iter_rows, start=2):
                    if r is None or not any(x is not None for x in r):
                        continue
                    row_dict: dict[str, Any] = {"_sheet": sheet_name, "_row": row_idx}
                    for i, h in enumerate(headers):
                        if h and i < len(r):
                            val = r[i]
                            if val is not None:
                                row_dict[h] = val
                    if len(row_dict) > 2:
                        rows.append(row_dict)
            wb.close()
            if not rows:
                return None, "Excel 数据源未提取到有效数据行"
            return rows, None
        except Exception as openpyxl_exc:
            try:
                import pandas as pd

                excel_file = pd.ExcelFile(source_path)
                clean_rows = []
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    records = df.to_dict(orient="records")
                    for row_idx, rec in enumerate(records, start=1):
                        clean_rec = {
                            str(k).strip(): v
                            for k, v in rec.items()
                            if pd.notna(v)
                        }
                        if clean_rec:
                            clean_rec["_sheet"] = str(sheet_name)
                            clean_rec["_row"] = row_idx
                            clean_rows.append(clean_rec)
                if not clean_rows:
                    return None, "Excel 数据源未提取到有效数据行"
                return clean_rows, None
            except Exception as pd_exc:
                return None, f"无法解析 Excel 数据源 ({openpyxl_exc}; {pd_exc})"

    if ext == ".xls":
        try:
            import pandas as pd

            excel_file = pd.ExcelFile(source_path)
            clean_rows = []
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                records = df.to_dict(orient="records")
                for row_idx, rec in enumerate(records, start=1):
                    clean_rec = {
                        str(k).strip(): v
                        for k, v in rec.items()
                        if pd.notna(v)
                    }
                    if clean_rec:
                        clean_rec["_sheet"] = str(sheet_name)
                        clean_rec["_row"] = row_idx
                        clean_rows.append(clean_rec)
            if not clean_rows:
                return None, "XLS 数据源未提取到有效数据行"
            return clean_rows, None
        except Exception as exc:
            return None, f"无法解析 XLS 数据源: {exc}"

    if ext == ".json":
        try:
            with open(source_path, encoding="utf-8", errors="replace") as f:
                data = json.load(f)

            flat_records: list[dict[str, Any]] = []

            def _traverse_json(node: Any, prefix: str = "") -> None:
                if isinstance(node, dict):
                    if "id" in node or "metric_id" in node or "name" in node:
                        flat_records.append(dict(node))
                    row_entry: dict[str, Any] = {}
                    for k, v in node.items():
                        full_k = f"{prefix}.{k}" if prefix else str(k)
                        if isinstance(v, (int, float, str, bool)) or v is None:
                            row_entry[full_k] = v
                        elif isinstance(v, (dict, list)):
                            _traverse_json(v, full_k)
                    if row_entry:
                        flat_records.append(row_entry)
                elif isinstance(node, list):
                    for idx, item in enumerate(node):
                        item_prefix = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
                        _traverse_json(item, item_prefix)

            _traverse_json(data)
            if not flat_records:
                return None, "JSON 数据源未解析出结构化键值"
            return flat_records, None
        except Exception as exc:
            return None, f"无法解析 JSON 数据源: {exc}"

    return None, f"unsupported_format:{ext}"


def _check_source_metrics_equivalence(
    source_path: str,
    metrics: list[dict[str, Any]],
) -> tuple[bool, str]:
    """核验数据源文件重新生成后，其关键数值是否与原 frozen_results 声明的指标保持等价（严格 fail-closed）。"""
    if not metrics:
        return False, "未指定可供核验的绑定指标"

    if not os.path.isfile(source_path):
        return False, f"数据源文件不存在: {os.path.basename(source_path)}"

    rows, err = _extract_rows_from_source(source_path)
    if err:
        return False, f"数据源解析失败: {err}"

    if not rows:
        return False, "数据源提取结果为空，无法进行等价性比对"

    for metric in metrics:
        metric_id = str(metric.get("id", "")).strip()
        metric_label = str(metric.get("label", "")).strip()
        expected_val = metric.get("value")

        if not metric_id:
            return False, "待核验指标缺少 id 字段"

        if (
            expected_val is None
            or isinstance(expected_val, bool)
            or not isinstance(expected_val, (int, float))
            or not math.isfinite(expected_val)
        ):
            return False, f"指标 '{metric_id}' 的期望值非有效有限数: {expected_val}"

        expected_float = float(expected_val)
        metric_id_lower = metric_id.lower()
        metric_label_lower = metric_label.lower() if metric_label else ""

        matches: list[tuple[str, float]] = []

        for row in rows:
            sheet_tag = f"[{row.get('_sheet')}]" if "_sheet" in row else ""
            row_tag = f"row {row.get('_row')}" if "_row" in row else ""
            loc = f"{sheet_tag} {row_tag}".strip() or "record"

            # 场景 1: 单行宽表或包含对应列名
            for col_name, val_str in row.items():
                if col_name.startswith("_") or val_str is None or isinstance(val_str, bool):
                    continue
                clean_col = col_name.strip().lower()
                matched = False
                if metric_id_lower == clean_col or clean_col.endswith(f".{metric_id_lower}"):
                    matched = True
                elif metric_label_lower and (metric_label_lower == clean_col or clean_col.endswith(f".{metric_label_lower}")):
                    # 仅在整行无任何显式匹配 ID 列且该列名精确匹配标签时允许兼容匹配
                    if not any(c.strip().lower() == metric_id_lower or c.strip().lower().endswith(f".{metric_id_lower}") for c in row if not c.startswith("_")):
                        matched = True

                if matched:
                    try:
                        parsed = float(val_str)
                        if not math.isfinite(parsed):
                            return False, f"指标 '{metric_id}' 在 {os.path.basename(source_path)} 中解析为非有限数值 (NaN/Infinity)"
                        matches.append((f"{loc}:{col_name}", parsed))
                    except (ValueError, TypeError):
                        return False, f"指标 '{metric_id}' 在 {os.path.basename(source_path)} 中的值无法解析为浮点数: {val_str!r}"

            # 场景 2: 键值对长表结构（如 指标ID, 指标名称, 数值）
            row_id = str(
                row.get("指标ID")
                or row.get("metric_id")
                or row.get("id")
                or row.get("metric")
                or ""
            ).strip().lower()
            row_label = str(
                row.get("指标名称")
                or row.get("name")
                or row.get("label")
                or ""
            ).strip().lower()

            long_matched = False
            if row_id:
                # 存在显式 ID 字段时，必须严格匹配 metric_id，绝不允许通过标签覆盖冲突的 ID
                if metric_id_lower == row_id:
                    long_matched = True
            elif metric_label_lower and metric_label_lower == row_label:
                # 仅在行中无任何 ID 字段时，才允许按标签别名匹配
                long_matched = True

            if long_matched:
                raw_val = row.get("数值") or row.get("value") or row.get("actual")
                if raw_val is not None and not isinstance(raw_val, bool):
                    try:
                        parsed = float(raw_val)
                        if not math.isfinite(parsed):
                            return False, f"指标 '{metric_id}' 在长表行中解析为非有限数值 (NaN/Infinity)"
                        matches.append((f"{loc}:long_row", parsed))
                    except (ValueError, TypeError):
                        return False, f"指标 '{metric_id}' 对应长表数值无法解析为浮点数: {raw_val!r}"

        # 校验唯一定位与数值等价
        if len(matches) == 0:
            return (
                False,
                f"指标 '{metric_id}' 在 {os.path.basename(source_path)} 中缺失或无法定位（可能已被删除或改名）",
            )

        if len(matches) > 1:
            locs = ", ".join(m[0] for m in matches[:5])
            return (
                False,
                f"指标 '{metric_id}' 在 {os.path.basename(source_path)} 中定位到 {len(matches)} 处匹配 ({locs})，存在多行/跨Sheet重复或一旧一新定义，无法唯一定位",
            )

        found_loc, found_val = matches[0]
        tol = max(1e-5, abs(expected_float) * 1e-4)
        if abs(found_val - expected_float) > tol:
            return (
                False,
                f"指标 '{metric_id}' 在 {os.path.basename(source_path)} ({found_loc}) 中的数值已变更为 {found_val}（原冻结值: {expected_val}），属于实质性数据变更，禁止静默刷新哈希！",
            )

    return True, "metrics_equivalent"


def refresh_frozen_results_hashes(
    work_dir: str,
    verify_equivalence: bool = True,
) -> dict[str, Any]:
    """扫描任务目录下已存在的数据源文件，在保证数值严格等价的前提下同步刷新 frozen_results.json 中的 SHA-256 哈希。

    若检测到关键指标缺失、多重冲突、数据源文件缺失、路径越界或实质性数值变化，一律 fail-closed，拒绝改写文件。
    """
    for relative_path in FREEZE_FILENAMES:
        path = os.path.join(work_dir, relative_path)
        if not os.path.isfile(path):
            continue
        try:
            with open(path, encoding="utf-8") as handle:
                document = json.load(handle)
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            continue
        if not isinstance(document, dict):
            continue

        sources = document.get("sources", [])
        metrics_list = document.get("metrics", [])
        if not isinstance(sources, list):
            continue

        updated_count = 0
        updated_sources = []
        conflicts = []

        for source in sources:
            if not isinstance(source, dict):
                continue
            rel_path = source.get("relative_path") or source.get("path")
            if not rel_path:
                continue
            source_path = _safe_path(work_dir, str(rel_path))
            if source_path is None:
                conflicts.append({
                    "path": rel_path,
                    "reason": "数据源路径越界或非法",
                })
                continue

            if not os.path.isfile(source_path):
                conflicts.append({
                    "path": rel_path,
                    "reason": f"数据源文件缺失或不存在: {rel_path}",
                })
                continue

            fresh_hash = _sha256(source_path)
            old_hash = source.get("sha256")
            if old_hash != fresh_hash:
                norm_rel = str(rel_path).replace("\\", "/").strip().lstrip("./")
                # 查找严格绑定到该相对路径的 metrics（严禁使用 basename 兜底）
                bound_metrics = [
                    m
                    for m in metrics_list
                    if isinstance(m, dict)
                    and (
                        str(m.get("source_path", "")).replace("\\", "/").strip().lstrip("./") == norm_rel
                        or str(m.get("source", "")).replace("\\", "/").strip().lstrip("./") == norm_rel
                    )
                ]

                if not bound_metrics:
                    conflicts.append({
                        "path": rel_path,
                        "old_sha256": old_hash,
                        "new_sha256": fresh_hash,
                        "reason": "未找到绑定指标，无法证明等价性",
                    })
                    continue

                is_equiv, reason = _check_source_metrics_equivalence(
                    source_path, bound_metrics
                )
                if not is_equiv:
                    conflicts.append({
                        "path": rel_path,
                        "old_sha256": old_hash,
                        "new_sha256": fresh_hash,
                        "reason": reason,
                    })
                    continue

                source["sha256"] = fresh_hash
                updated_count += 1
                updated_sources.append({
                    "path": rel_path,
                    "old_sha256": old_hash,
                    "new_sha256": fresh_hash,
                    "status": "EQUIVALENT_REFRESH",
                })

        if updated_count > 0 and not conflicts:
            try:
                with open(path, "w", encoding="utf-8") as handle:
                    json.dump(document, handle, ensure_ascii=False, indent=2)
            except OSError:
                return {
                    "active": True,
                    "updated": False,
                    "updated_count": 0,
                    "error": f"写入 {relative_path} 失败",
                    "conflicts": conflicts,
                    "has_conflicts": True,
                }

        return {
            "active": True,
            "updated": updated_count > 0 and not conflicts,
            "updated_count": updated_count if not conflicts else 0,
            "path": relative_path,
            "updated_sources": updated_sources if not conflicts else [],
            "conflicts": conflicts,
            "has_conflicts": len(conflicts) > 0,
        }

    return {"active": False, "updated": False, "updated_count": 0, "conflicts": [], "has_conflicts": False}
